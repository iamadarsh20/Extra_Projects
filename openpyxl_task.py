# from openpyxl import load_workbook

# workbook_name = 'sample.xlsx'
# wb = load_workbook(workbook_name)
# page = wb.active

# # New data to write:
# new_companies = [['name3','address3','tel3','web3'], ['name4','address4','tel4','web4']]

# for info in new_companies:
#     page.append(info)


# wb.save(r"E:\sequel_task\temp.xlsx")

# from openpyxl import Workbook  
  
# wb = Workbook()  
# sheet = wb.active  
  
# data = (  
#     (11, 48, 50),  
#     (81, 30, 82),  
#     (20, 51, 72),  
#     (21, 14, 60),  
#     (28, 41, 49),  
#     (74, 65, 53),  
#     ("Peter", 'Andrew',45.63)  
# )  
  
# for i in data:
#     if i%2!=0:
#         sheet.append(i)  
      
# wb.save(r"E:\sequel_task\temp.xlsx")  











import openpyxl
workbook = openpyxl.Workbook()
sheet = workbook.active

a = ['adarsh', 'Ram', 'Shyam','shrey', 'abhishek', 'raju','harish', 'rosy', 'sati']
# c = ['shrey', 'abhishek', 'raju']
# e = ['harish', 'rosy', 'sati']
# g = ['Anand', 'Yash','rakesh']
# i = [121, 768, 890]


# for k in range(1,len(a)+1):

#     print(a[k-1],c[k-1],e[k-1],g[k-1],i[k-1])
#     sheet[f'A{k}'].value = a[k-1]
#     sheet[f'C{k}'].value = c[k-1]
#     sheet[f'E{k}'].value = e[k-1]
#     sheet[f'G{k}'].value = g[k-1]
#     sheet[f'I{k}'].value = i[k-1]
#     # print()


sh = []
for i in range(65,70):
    for j in range(1,6):
        if (i & j)%2!=0:
            cell = chr(i)+str(j)
            sh.append(cell)
            # print(cell,end=" ")
print(sh)

for i in range(len(a)):
    # print(i)

    sheet[sh] = a[i]



            # sheet[cell]=1
            # for k in range(9):
            #     sheet[f'cell{k}'].value = a[k]
                
               
        
#     print()

# workbook.save(r"E:\sequel_task\temp.xlsx")






# # for j in range(len(a)):

   
# #     for i in range(1,10):
# #         if i%2!=0:
# #             sheet[f'A{i}'].value = a[j]
# #             # sheet[f'B{i}'].value = c[k]
# #             # sheet[f'C{i}'].value = e[l]
# #             workbook.save(r"E:\sequel_task\temp.xlsx")

            
            
            
